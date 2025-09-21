import csv
import io
import os
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook


# Проверка, что архив был создан

def test_create_zip_file(create_zip):
    assert os.path.isfile(create_zip)
    assert create_zip.endswith("ZIP_example.zip")

# Проверка, что список файлов в ZIP совпадает с ожидаемым

def test_files_list_in_zip(create_zip):
    with zipfile.ZipFile(create_zip, mode='r') as zip_file:
        actual_files = zip_file.namelist()

    assert actual_files == ['CSV_Example.csv', 'PDF_Example.pdf', 'XLSX_Example.xlsx']

# Проверка размера файла, количества строк, столбцов и названий заголовков в CSV файле

def test_csv_file(create_zip):
    with zipfile.ZipFile(create_zip, mode='r') as zip_file:
        csv_size = zip_file.getinfo('CSV_Example.csv').file_size
        with zip_file.open("CSV_Example.csv") as csv_file:
            # Получение текстового файла. Нужно, чтобы не падал с ошибкой next
            text_file = io.TextIOWrapper(csv_file, encoding='utf-8')
            reader = csv.reader(text_file, delimiter=',')

            row_count = sum(1 for row in csv_file)
            text_file.seek(0) # Возвращение к началу файла, чтобы прочитать его заново

            headers = next(reader)

            column_count = len(headers)

    assert csv_size == 1856702
    assert row_count == 12869
    assert column_count == 8
    assert headers == ['Series_reference', 'Period', 'Data_value', 'STATUS', 'UNITS', 'Subject', 'Group',
                       'Series_title_1']


# Проверка размера, количества строк, столбцов, первого имени и последнего id в XLSX файле

def test_xlsx_file(create_zip):
    with zipfile.ZipFile(create_zip, mode='r') as zip_file:
        xlsx_size = zip_file.getinfo('XLSX_Example.xlsx').file_size
        with zip_file.open("XLSX_Example.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook['Sheet1']

            max_row = sheet.max_row
            max_column = sheet.max_column
            first_name = sheet.cell(row=2, column=2).value
            last_id = sheet.cell(row=max_row, column=max_column).value

    assert xlsx_size == 188887
    assert max_row == 5001
    assert max_column == 8
    assert first_name == 'Dulce'
    assert last_id == 6125

# Проверка размера файла, количества страниц, текста на первой и последней странице в PDF файле

def test_pdf_file(create_zip):
    with zipfile.ZipFile(create_zip, mode='r') as zip_file:
        pdf_size = zip_file.getinfo('PDF_Example.pdf').file_size
        with zip_file.open("PDF_Example.pdf") as pdf_file:
            reader = PdfReader(pdf_file)

            actual_page_count = len(reader.pages)
            first_page = reader.pages[0].extract_text().replace('\n', '')
            last_page = reader.pages[-1].extract_text().replace('\n', '')


    assert pdf_size == 469513
    assert actual_page_count == 5
    assert "Lorem ipsum dolor sit amet, consectetur adipiscing elit. Nunc ac faucibus odio" in first_page
    assert "Ut ac loremsed turpis imperdiet eleifend sit amet id sapien" in last_page
